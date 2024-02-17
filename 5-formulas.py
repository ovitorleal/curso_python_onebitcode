from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#1 le pasta de trabalho e planilha
wb = load_workbook("data/barchart.xlsx")
sheet = wb["Relatório"]

#2 Referencias das linhas e colunas
min_column = wb.active.min_column
max_coluum = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#3 incluindo formulas
# sheet["B6"] = "=SUM(B2:B5)"
# sheet["B6"].style = "Currency" #forma manual.. teria que repetir 7 vezes.
for i in range(min_column+1, max_coluum+1): #+1 pq nao quero os titulos 
    letter = get_column_letter(i)
    sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row})"
    sheet[f"{letter}{max_row+1}"].style = "Currency"
    
wb.save("test.xlsx")


