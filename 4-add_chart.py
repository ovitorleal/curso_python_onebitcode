from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

#1 le pasta de trabalho e planilha
wb = load_workbook("data/pivot_table.xlsx")
sheet = wb["Relatório"]

#2 Referencias das linhas e colunas
min_column = wb.active.min_column
max_coluum = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#3 adicionando dados e categorias do grafico
barchart = BarChart()

data = Reference(
    sheet,
    min_col=min_column + 1,
    max_col= max_coluum,
    min_row=min_row,
    max_row=max_row
)

categories = Reference(
    sheet,
    min_col=min_column,
    max_col=min_column, #min_column ao inves de max pq é apenas a parte de referencias para o grafico ficar melhor formatado
    min_row=min_row+1,
    max_row=max_row
)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

#4 criando o grafico propriamente dito
sheet.add_chart(barchart, "B10")
barchart.title= "Vendas por Fabricante"
barchart.style = 2


#5 salvando o workboob
wb.save("data/barchart.xlsx")
